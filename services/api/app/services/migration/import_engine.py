"""Import engine — processes migration jobs from detected platform data.

Import order: companies -> contacts -> candidates -> jobs -> placements -> notes.
Batched commits (100 rows), dedup by email/name, parent-child resolution.
Creates real Winnow entities (User, CandidateProfile, EmployerProfile, EmployerJob).
"""

import csv
import logging
from datetime import UTC, datetime
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.candidate_profile import CandidateProfile
from app.models.employer import EmployerProfile
from app.models.migration import MigrationEntityMap, MigrationJob
from app.models.user import User
from app.services.migration.platform_detector import FIELD_MAPPINGS, detect_platform

logger = logging.getLogger(__name__)

BATCH_SIZE = 100


def run_migration(migration_job_id: int, db: Session) -> dict:
    """Execute a migration job — main pipeline.

    Steps:
    1. Load migration job config
    2. Detect platform (if not already set)
    3. Read CSV and map fields
    4. Import entities, creating real Winnow records
    5. Resolve parent-child relationships
    6. Update job stats
    """
    job = db.execute(
        select(MigrationJob).where(MigrationJob.id == migration_job_id)
    ).scalar_one_or_none()
    if not job:
        raise ValueError(f"Migration job {migration_job_id} not found")

    job.status = "importing"
    job.started_at = datetime.now(UTC)
    db.commit()

    stats = {"imported": 0, "merged": 0, "skipped": 0, "errors": 0, "by_type": {}}
    errors: list[dict] = []

    try:
        file_path = job.source_file_path
        if not file_path or not Path(file_path).exists():
            raise FileNotFoundError(f"Source file not found: {file_path}")

        # Detect platform if needed
        if not job.source_platform_detected:
            detection = detect_platform(file_path)
            job.source_platform_detected = detection["platform"]
            db.commit()

        platform = job.source_platform_detected or job.source_platform
        field_mapping = FIELD_MAPPINGS.get(platform, {})

        # Determine entity type from config or default
        config = job.config_json or {}
        entity_type = config.get("entity_type", "candidates")

        # Read and import
        rows = _read_csv(file_path)
        type_stats = {"imported": 0, "merged": 0, "skipped": 0, "errors": 0}

        for i, row in enumerate(rows):
            try:
                mapped = _map_row(row, field_mapping)
                source_id = mapped.get("external_id") or mapped.get("email") or str(i)

                if entity_type == "candidates":
                    result = _import_candidate(
                        migration_job_id, source_id, mapped, row, db
                    )
                elif entity_type == "companies":
                    result = _import_company(
                        migration_job_id, source_id, mapped, row, db
                    )
                elif entity_type in ("jobs", "contacts", "placements", "notes"):
                    result = _import_generic(
                        migration_job_id, entity_type, source_id, mapped, row, db
                    )
                else:
                    result = _import_generic(
                        migration_job_id, entity_type, source_id, mapped, row, db
                    )

                type_stats[result] += 1
                stats[result] += 1

                # Batch commit every BATCH_SIZE rows
                if (i + 1) % BATCH_SIZE == 0:
                    db.commit()

            except Exception as e:
                type_stats["errors"] += 1
                stats["errors"] += 1
                errors.append({"row": i, "error": str(e)})
                logger.warning("Row %d import error: %s", i, e)

        db.commit()
        stats["by_type"][entity_type] = type_stats

        # Resolve parent-child relationships
        _resolve_relationships(migration_job_id, db)

        job.status = "completed"
        job.completed_at = datetime.now(UTC)

    except Exception as e:
        job.status = "failed"
        errors.append({"error": str(e), "fatal": True})
        logger.exception("Migration job %d failed", migration_job_id)

    job.stats_json = stats
    job.error_log = errors if errors else None
    job.updated_at = datetime.now(UTC)
    db.commit()

    return {"job_id": migration_job_id, "status": job.status, "stats": stats}


# ---------------------------------------------------------------------------
# Entity import functions
# ---------------------------------------------------------------------------


def _import_candidate(
    migration_job_id: int,
    source_id: str,
    mapped: dict,
    raw: dict,
    db: Session,
) -> str:
    """Import a candidate row — creates User + CandidateProfile.

    Returns: "imported" | "merged" | "skipped"
    """
    email = mapped.get("email", "").strip().lower()

    # Dedup by email across entire database
    existing_user = None
    existing_profile = None
    if email:
        existing_user = db.execute(
            select(User).where(User.email == email)
        ).scalar_one_or_none()
        if existing_user:
            existing_profile = db.execute(
                select(CandidateProfile).where(
                    CandidateProfile.user_id == existing_user.id
                )
            ).scalar_one_or_none()

    # Build profile_json from mapped data
    profile_json = _build_profile_json(mapped)

    if existing_profile:
        # Merge: update profile_json without overwriting user-edited fields
        current = existing_profile.profile_json or {}
        for key, val in profile_json.items():
            if val and not current.get(key):
                current[key] = val
        # Always update skills if new ones found
        if profile_json.get("skills"):
            existing_skills = set(current.get("skills", []))
            new_skills = set(profile_json.get("skills", []))
            current["skills"] = list(existing_skills | new_skills)
        existing_profile.profile_json = current
        winnow_id = existing_profile.id

        _record_entity(
            migration_job_id,
            "candidate",
            source_id,
            "candidate_profile",
            winnow_id,
            raw,
            "merged",
            db,
        )
        return "merged"

    # Create new User + CandidateProfile
    if not email:
        email = f"migrated-{source_id}@imported.winnow"

    user = User(email=email, password_hash="", role="candidate")
    db.add(user)
    db.flush()

    profile = CandidateProfile(
        user_id=user.id,
        version=1,
        profile_json=profile_json,
    )
    db.add(profile)
    db.flush()

    _record_entity(
        migration_job_id,
        "candidate",
        source_id,
        "candidate_profile",
        profile.id,
        raw,
        "imported",
        db,
    )
    return "imported"


def _import_company(
    migration_job_id: int,
    source_id: str,
    mapped: dict,
    raw: dict,
    db: Session,
) -> str:
    """Import a company row — creates User + EmployerProfile.

    Returns: "imported" | "merged" | "skipped"
    """
    company_name = (mapped.get("company_name") or mapped.get("company") or "").strip()
    if not company_name:
        _record_entity(
            migration_job_id,
            "company",
            source_id,
            "employer_profile",
            None,
            raw,
            "skipped",
            db,
        )
        return "skipped"

    # Dedup by company name
    existing = db.execute(
        select(EmployerProfile).where(EmployerProfile.company_name == company_name)
    ).scalar_one_or_none()

    if existing:
        # Merge non-empty fields
        if mapped.get("website") and not existing.company_website:
            existing.company_website = mapped["website"]
        if mapped.get("industry") and not existing.industry:
            existing.industry = mapped["industry"]

        _record_entity(
            migration_job_id,
            "company",
            source_id,
            "employer_profile",
            existing.id,
            raw,
            "merged",
            db,
        )
        return "merged"

    # Create new employer user + profile
    email = mapped.get("email") or f"company-{source_id}@imported.winnow"
    user = User(email=email, password_hash="", role="employer")
    db.add(user)
    db.flush()

    employer = EmployerProfile(
        user_id=user.id,
        company_name=company_name,
        company_website=mapped.get("website"),
        industry=mapped.get("industry"),
    )
    db.add(employer)
    db.flush()

    _record_entity(
        migration_job_id,
        "company",
        source_id,
        "employer_profile",
        employer.id,
        raw,
        "imported",
        db,
    )
    return "imported"


def _import_generic(
    migration_job_id: int,
    entity_type: str,
    source_id: str,
    mapped: dict,
    raw: dict,
    db: Session,
) -> str:
    """Import a generic entity type (jobs, contacts, placements, notes).

    Stores mapped data in entity_map for later relationship resolution.
    """
    winnow_type = _winnow_type(entity_type)

    # Check for duplicate within this migration
    existing = db.execute(
        select(MigrationEntityMap).where(
            MigrationEntityMap.migration_job_id == migration_job_id,
            MigrationEntityMap.source_entity_type == entity_type,
            MigrationEntityMap.source_entity_id == str(source_id),
        )
    ).scalar_one_or_none()

    if existing:
        return "skipped"

    # Determine parent source ID for relationship linking
    parent_id = (
        mapped.get("company_source_id") or mapped.get("candidate_source_id") or None
    )

    entity = MigrationEntityMap(
        migration_job_id=migration_job_id,
        source_entity_type=entity_type,
        source_entity_id=str(source_id),
        winnow_entity_type=winnow_type,
        parent_source_id=parent_id,
        raw_data=mapped,
        status="imported",
    )
    db.add(entity)

    return "imported"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _build_profile_json(mapped: dict) -> dict:
    """Convert flat mapped fields into a structured profile_json."""
    profile: dict = {}

    # Handle dot-notation keys (e.g., basics.first_name)
    for key, val in mapped.items():
        if "." in key:
            parts = key.split(".", 1)
            parent = parts[0]
            child = parts[1]
            if parent not in profile:
                profile[parent] = {}
            profile[parent][child] = val
        else:
            profile[key] = val

    # Build canonical structure from flat keys
    basics: dict = profile.get("basics", {})
    if mapped.get("first_name"):
        basics["first_name"] = mapped["first_name"]
    if mapped.get("last_name"):
        basics["last_name"] = mapped["last_name"]
    if mapped.get("email"):
        basics["email"] = mapped["email"]
    if mapped.get("phone"):
        basics["phone"] = mapped["phone"]
    if mapped.get("linkedin_url"):
        basics["linkedin_url"] = mapped["linkedin_url"]
    if mapped.get("company"):
        basics["current_company"] = mapped["company"]
    if mapped.get("job_title"):
        basics["current_title"] = mapped["job_title"]
    if mapped.get("city") or mapped.get("state"):
        parts = [mapped.get("city", ""), mapped.get("state", "")]
        basics["location"] = ", ".join(p for p in parts if p)

    if basics:
        profile["basics"] = basics

    # Name
    name_parts = [mapped.get("first_name", ""), mapped.get("last_name", "")]
    name = " ".join(p for p in name_parts if p)
    if name:
        profile["name"] = name

    # Skills — split comma-separated string into list
    skills_raw = mapped.get("skills", "")
    if isinstance(skills_raw, str) and skills_raw:
        profile["skills"] = [s.strip() for s in skills_raw.split(",") if s.strip()]
    elif isinstance(skills_raw, list):
        profile["skills"] = skills_raw

    # Custom fields — preserve anything not in standard mapping
    standard_keys = {
        "first_name",
        "last_name",
        "email",
        "phone",
        "company",
        "job_title",
        "skills",
        "external_id",
        "status",
        "source",
        "owner",
        "created_at",
        "salary",
        "linkedin_url",
        "city",
        "state",
        "location",
        "employment_type",
        "salary_min",
        "salary_max",
        "tags",
        "website",
        "industry",
        "company_name",
        "company_source_id",
        "candidate_source_id",
    }
    custom = {
        k: v for k, v in mapped.items() if k not in standard_keys and "." not in k
    }
    if custom:
        profile["custom_fields"] = custom

    # Salary
    if mapped.get("salary"):
        profile.setdefault("preferences", {})["salary_current"] = mapped["salary"]

    # Tags
    if mapped.get("tags"):
        tags = mapped["tags"]
        if isinstance(tags, str):
            profile["tags"] = [t.strip() for t in tags.split(",") if t.strip()]
        else:
            profile["tags"] = tags

    # Source tracking
    if mapped.get("source"):
        profile["source"] = mapped["source"]

    return profile


def _record_entity(
    migration_job_id: int,
    source_type: str,
    source_id: str,
    winnow_type: str,
    winnow_id: int | None,
    raw: dict,
    status: str,
    db: Session,
) -> None:
    """Record an entity mapping in migration_entity_map."""
    entity = MigrationEntityMap(
        migration_job_id=migration_job_id,
        source_entity_type=source_type,
        source_entity_id=str(source_id),
        winnow_entity_type=winnow_type,
        winnow_entity_id=winnow_id,
        raw_data=raw,
        status=status,
    )
    db.add(entity)


def _read_csv(file_path: str) -> list[dict]:
    """Read CSV or XLSX into list of dicts."""
    if file_path.lower().endswith(".xlsx"):
        return _read_xlsx(file_path)
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader)


def _read_xlsx(file_path: str) -> list[dict]:
    """Read XLSX file into list of dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, ())
    headers = [
        str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)
    ]
    result = []
    for row in rows:
        record = {}
        for h, val in zip(headers, row, strict=False):
            record[h] = str(val).strip() if val is not None else ""
        result.append(record)
    wb.close()
    return result


def _map_row(row: dict, field_mapping: dict) -> dict:
    """Map source columns to canonical Winnow field names."""
    mapped = {}
    for source_col, value in row.items():
        canonical = field_mapping.get(source_col, source_col)
        if value and value.strip():
            mapped[canonical] = value.strip()
    return mapped


def _winnow_type(source_type: str) -> str:
    """Map source entity type to Winnow entity type."""
    mapping = {
        "candidates": "candidate_profile",
        "jobs": "employer_job",
        "companies": "employer_profile",
        "contacts": "user",
        "placements": "match",
        "notes": "note",
    }
    return mapping.get(source_type, source_type)


def _resolve_relationships(migration_job_id: int, db: Session) -> None:
    """Link parent-child entities using parent_source_id references."""
    entities = (
        db.execute(
            select(MigrationEntityMap).where(
                MigrationEntityMap.migration_job_id == migration_job_id,
                MigrationEntityMap.parent_source_id.isnot(None),
            )
        )
        .scalars()
        .all()
    )

    for entity in entities:
        parent = db.execute(
            select(MigrationEntityMap).where(
                MigrationEntityMap.migration_job_id == migration_job_id,
                MigrationEntityMap.source_entity_id == entity.parent_source_id,
            )
        ).scalar_one_or_none()

        if parent and parent.winnow_entity_id:
            data = entity.raw_data or {}
            data["_parent_winnow_id"] = parent.winnow_entity_id
            entity.raw_data = data

    db.commit()


def get_preview(migration_job_id: int, db: Session, limit: int = 10) -> dict:
    """Preview imported entities for a migration job."""
    entities = (
        db.execute(
            select(MigrationEntityMap)
            .where(MigrationEntityMap.migration_job_id == migration_job_id)
            .limit(limit)
        )
        .scalars()
        .all()
    )

    return {
        "job_id": migration_job_id,
        "preview_count": len(entities),
        "entities": [
            {
                "id": e.id,
                "source_type": e.source_entity_type,
                "source_id": e.source_entity_id,
                "winnow_type": e.winnow_entity_type,
                "winnow_id": e.winnow_entity_id,
                "status": e.status,
                "data": e.raw_data,
            }
            for e in entities
        ],
    }


def rollback_migration(migration_job_id: int, db: Session) -> dict:
    """Delete all entities created by a migration job.

    Removes entity_map records and the actual Winnow entities they reference.
    """
    job = db.execute(
        select(MigrationJob).where(MigrationJob.id == migration_job_id)
    ).scalar_one_or_none()
    if not job:
        raise ValueError("Migration job not found")

    # Load all entity mappings with winnow IDs
    mappings = (
        db.execute(
            select(MigrationEntityMap).where(
                MigrationEntityMap.migration_job_id == migration_job_id,
                MigrationEntityMap.winnow_entity_id.isnot(None),
                MigrationEntityMap.status.in_(["imported"]),
            )
        )
        .scalars()
        .all()
    )

    # Delete actual Winnow entities (reverse order to handle FK deps)
    deleted_profiles = 0
    deleted_users = set()
    for m in mappings:
        if m.winnow_entity_type == "candidate_profile":
            profile = db.get(CandidateProfile, m.winnow_entity_id)
            if profile:
                if profile.user_id:
                    deleted_users.add(profile.user_id)
                db.delete(profile)
                deleted_profiles += 1
        elif m.winnow_entity_type == "employer_profile":
            employer = db.get(EmployerProfile, m.winnow_entity_id)
            if employer:
                deleted_users.add(employer.user_id)
                db.delete(employer)

    # Delete placeholder users created during import
    for uid in deleted_users:
        user = db.get(User, uid)
        if user and (
            user.email.endswith("@imported.winnow")
            or user.email.endswith("@sourced.winnow")
        ):
            db.delete(user)

    # Delete entity map records
    count = (
        db.query(MigrationEntityMap)
        .filter(MigrationEntityMap.migration_job_id == migration_job_id)
        .delete()
    )

    job.status = "rolled_back"
    job.updated_at = datetime.now(UTC)
    db.commit()

    return {
        "job_id": migration_job_id,
        "entity_maps_deleted": count,
        "profiles_deleted": deleted_profiles,
        "status": "rolled_back",
    }

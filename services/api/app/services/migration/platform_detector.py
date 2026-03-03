"""Platform fingerprinting — detect which ATS/CRM an export came from.

Supports CSV, JSON, and ZIP files from:
Bullhorn, Recruit CRM, CATSOne, Zoho Recruit, generic fallback.
Includes AI-assisted mapping for unknown formats.
"""

import csv
import json as json_mod
import logging
import os
import zipfile
from pathlib import Path

logger = logging.getLogger(__name__)

# Platform header signatures.
# Maps platform -> entity_type -> required + optional headers.
PLATFORM_SIGNATURES: dict[str, dict[str, dict]] = {
    "bullhorn": {
        "candidates": {
            "required": [
                "firstName",
                "lastName",
                "email",
                "status",
                "source",
            ],
            "optional": [
                "candidateID",
                "dateAdded",
                "owner",
                "skillList",
                "salary",
                "companyName",
                "customText1",
            ],
        },
        "jobs": {
            "required": ["title", "clientCorporation", "status", "dateAdded"],
            "optional": [
                "jobOrderID",
                "employmentType",
                "salary",
                "startDate",
                "owner",
                "numOpenings",
            ],
        },
        "companies": {
            "required": ["companyName", "status"],
            "optional": ["clientCorporationID", "phone", "address", "industry"],
        },
        "placements": {
            "required": ["candidate", "jobOrder", "status"],
            "optional": ["placementID", "salary", "startDate", "fee"],
        },
    },
    "recruitcrm": {
        "candidates": {
            "required": ["First Name", "Last Name", "Email"],
            "optional": [
                "Phone",
                "Current Company",
                "Current Title",
                "Source",
                "Tags",
                "Resume",
                "Created At",
            ],
        },
        "jobs": {
            "required": ["Job Title", "Company", "Status"],
            "optional": [
                "Location",
                "Job Type",
                "Salary Min",
                "Salary Max",
                "Created At",
            ],
        },
        "contacts": {
            "required": ["First Name", "Last Name", "Email", "Company"],
            "optional": ["Phone", "Title", "Tags"],
        },
    },
    "catsone": {
        "candidates": {
            "required": ["First Name", "Last Name", "E-Mail"],
            "optional": [
                "ID",
                "Phone",
                "City",
                "State",
                "Source",
                "Date Created",
                "Owner",
            ],
        },
        "jobs": {
            "required": ["Title", "Company", "Status"],
            "optional": ["ID", "City", "State", "Type", "Date Created", "Recruiter"],
        },
    },
    "zoho_recruit": {
        "candidates": {
            "required": ["First Name", "Last Name", "Email"],
            "optional": [
                "Phone",
                "Current Job Title",
                "Current Employer",
                "Candidate Status",
                "Source",
                "Created Time",
            ],
        },
        "jobs": {
            "required": ["Job Opening Name", "Client Name", "Job Opening Status"],
            "optional": [
                "City",
                "State",
                "Job Type",
                "Number of Positions",
                "Created Time",
            ],
        },
    },
}

# Filename hints
FILENAME_HINTS: dict[str, list[str]] = {
    "bullhorn": ["bullhorn", "bh_export", "bh-export"],
    "recruitcrm": ["recruitcrm", "rcrm", "recruit_crm"],
    "catsone": ["catsone", "cats_", "cats-"],
    "zoho_recruit": ["zoho", "zr_export", "zoho_recruit"],
}

# Field mappings per platform (source column -> Winnow canonical name)
FIELD_MAPPINGS: dict[str, dict[str, str]] = {
    "bullhorn": {
        "firstName": "first_name",
        "lastName": "last_name",
        "email": "email",
        "status": "status",
        "source": "source",
        "candidateID": "external_id",
        "dateAdded": "created_at",
        "owner": "owner",
        "skillList": "skills",
        "salary": "salary",
        "companyName": "company",
        "title": "job_title",
        "clientCorporation": "company",
        "jobOrderID": "external_id",
        "employmentType": "employment_type",
    },
    "recruitcrm": {
        "First Name": "first_name",
        "Last Name": "last_name",
        "Email": "email",
        "Phone": "phone",
        "Current Company": "company",
        "Current Title": "job_title",
        "Source": "source",
        "Tags": "tags",
        "Job Title": "job_title",
        "Company": "company",
        "Status": "status",
        "Location": "location",
        "Job Type": "employment_type",
        "Salary Min": "salary_min",
        "Salary Max": "salary_max",
    },
    "catsone": {
        "First Name": "first_name",
        "Last Name": "last_name",
        "E-Mail": "email",
        "ID": "external_id",
        "Phone": "phone",
        "City": "city",
        "State": "state",
        "Source": "source",
        "Date Created": "created_at",
        "Title": "job_title",
        "Company": "company",
        "Status": "status",
        "Type": "employment_type",
    },
    "zoho_recruit": {
        "First Name": "first_name",
        "Last Name": "last_name",
        "Email": "email",
        "Phone": "phone",
        "Current Job Title": "job_title",
        "Current Employer": "company",
        "Candidate Status": "status",
        "Source": "source",
        "Job Opening Name": "job_title",
        "Client Name": "company",
        "Job Opening Status": "status",
        "City": "city",
        "State": "state",
        "Job Type": "employment_type",
    },
}


def detect_platform(file_path: str) -> dict:
    """Detect which ATS platform produced an export file.

    Supports CSV, JSON, and ZIP files.

    Returns:
        {
            platform: str,
            confidence: float (0-1),
            evidence: list[str],
            entity_types_found: list[str],
            row_count: int,
            field_mapping: dict,
        }
    """
    path = Path(file_path)
    filename_lower = path.name.lower()
    suffix = path.suffix.lower()

    # Route to format-specific detection
    if suffix == ".json":
        return _detect_from_json(path, filename_lower)
    if suffix == ".zip":
        return _detect_from_zip(path, filename_lower)
    if suffix == ".xlsx":
        return _detect_from_xlsx(path, filename_lower)
    # Default: CSV
    return _detect_from_csv(path, filename_lower)


def _detect_from_csv(path: Path, filename_lower: str) -> dict:
    """Detect platform from CSV headers."""
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            row_count = sum(1 for _ in reader)
    except Exception as e:
        logger.warning("Failed to read CSV %s: %s", path, e)
        return {
            "platform": "unknown",
            "confidence": 0.0,
            "evidence": [f"Read error: {e}"],
            "entity_types_found": [],
            "row_count": 0,
            "field_mapping": {},
        }

    return _score_headers(headers, row_count, filename_lower)


def _detect_from_xlsx(path: Path, filename_lower: str) -> dict:
    """Detect platform from XLSX (Excel) headers."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, ())
        headers = [str(h) for h in header_row if h is not None]
        row_count = sum(1 for _ in rows)
        wb.close()
    except Exception as e:
        logger.warning("Failed to read XLSX %s: %s", path, e)
        return {
            "platform": "unknown",
            "confidence": 0.0,
            "evidence": [f"XLSX read error: {e}"],
            "entity_types_found": [],
            "row_count": 0,
            "field_mapping": {},
        }

    return _score_headers(headers, row_count, filename_lower)


def _detect_from_json(path: Path, filename_lower: str) -> dict:
    """Detect platform from JSON file structure."""
    try:
        with open(path, encoding="utf-8") as f:
            data = json_mod.load(f)
    except Exception as e:
        logger.warning("Failed to read JSON %s: %s", path, e)
        return {
            "platform": "unknown",
            "confidence": 0.0,
            "evidence": [f"JSON read error: {e}"],
            "entity_types_found": [],
            "row_count": 0,
            "field_mapping": {},
        }

    # JSON could be a list of records or a dict with entity arrays
    if isinstance(data, list) and data:
        headers = list(data[0].keys()) if isinstance(data[0], dict) else []
        row_count = len(data)
    elif isinstance(data, dict):
        # Check for entity-keyed structure (e.g., {"candidates": [...], "jobs": [...]})
        headers = []
        row_count = 0
        for _key, val in data.items():
            if isinstance(val, list) and val and isinstance(val[0], dict):
                headers.extend(val[0].keys())
                row_count += len(val)
        headers = list(set(headers))
    else:
        return {
            "platform": "generic",
            "confidence": 0.1,
            "evidence": ["Unrecognized JSON structure"],
            "entity_types_found": [],
            "row_count": 0,
            "field_mapping": {},
        }

    return _score_headers(headers, row_count, filename_lower)


def _detect_from_zip(path: Path, filename_lower: str) -> dict:
    """Detect platform from ZIP archive contents."""
    try:
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()
    except Exception as e:
        logger.warning("Failed to read ZIP %s: %s", path, e)
        return {
            "platform": "unknown",
            "confidence": 0.0,
            "evidence": [f"ZIP read error: {e}"],
            "entity_types_found": [],
            "row_count": 0,
            "field_mapping": {},
        }

    names_lower = [n.lower() for n in names]
    evidence: list[str] = []
    entity_types: list[str] = []

    # Check for Bullhorn multi-file export
    bullhorn_files = ["candidate", "clientcorporation", "joborder", "placement"]
    bh_matches = sum(1 for bf in bullhorn_files if any(bf in n for n in names_lower))
    if bh_matches >= 2:
        evidence.append(f"ZIP contains {bh_matches} Bullhorn entity files")
        for bf in bullhorn_files:
            if any(bf in n for n in names_lower):
                entity_types.append(
                    {
                        "candidate": "candidates",
                        "clientcorporation": "companies",
                        "joborder": "jobs",
                        "placement": "placements",
                    }.get(bf, bf)
                )
        return {
            "platform": "bullhorn",
            "confidence": min(0.3 + bh_matches * 0.2, 1.0),
            "evidence": evidence,
            "entity_types_found": entity_types,
            "row_count": 0,
            "field_mapping": FIELD_MAPPINGS.get("bullhorn", {}),
        }

    # Check for resume archive (ZIP of PDF/DOCX files)
    _resume_exts = (".pdf", ".docx", ".doc")
    resume_files = [
        n
        for n in names
        if any(n.lower().endswith(ext) for ext in _resume_exts)
        and "__MACOSX" not in n
        and not n.split("/")[-1].startswith(".")
    ]
    if resume_files:
        evidence.append(f"ZIP contains {len(resume_files)} resume files (PDF/DOCX)")
        return {
            "platform": "resume_archive",
            "confidence": 0.9,
            "evidence": evidence,
            "entity_types_found": ["resumes"],
            "row_count": len(resume_files),
            "field_mapping": {},
        }

    # Generic ZIP — check for CSV files inside
    csv_files = [n for n in names if n.lower().endswith(".csv")]
    if csv_files:
        evidence.append(f"ZIP contains {len(csv_files)} CSV files")
        return {
            "platform": "generic",
            "confidence": 0.3,
            "evidence": evidence,
            "entity_types_found": ["unknown"],
            "row_count": 0,
            "field_mapping": {},
        }

    return {
        "platform": "unknown",
        "confidence": 0.0,
        "evidence": ["ZIP contains no recognizable data files"],
        "entity_types_found": [],
        "row_count": 0,
        "field_mapping": {},
    }


def _score_headers(headers: list[str], row_count: int, filename_lower: str) -> dict:
    """Score headers against all platform signatures."""
    headers_set = set(headers)
    best_platform = "generic"
    best_score = 0.0
    best_evidence: list[str] = []
    best_entity_types: list[str] = []

    for platform, entity_types in PLATFORM_SIGNATURES.items():
        platform_score = 0.0
        evidence: list[str] = []
        found_types: list[str] = []

        # Filename hint: +0.2
        for hint in FILENAME_HINTS.get(platform, []):
            if hint in filename_lower:
                platform_score += 0.2
                evidence.append(f"Filename contains '{hint}'")
                break

        # Check each entity type
        for entity_type, sig in entity_types.items():
            required = sig["required"]
            optional = sig.get("optional", [])

            required_match = sum(1 for h in required if h in headers_set)
            optional_match = sum(1 for h in optional if h in headers_set)

            if required_match == len(required):
                platform_score += 0.5
                evidence.append(
                    f"All {len(required)} required headers match for {entity_type}"
                )
                found_types.append(entity_type)

                if optional:
                    opt_ratio = optional_match / len(optional)
                    platform_score += 0.3 * opt_ratio
                    if optional_match > 0:
                        evidence.append(
                            f"{optional_match}/{len(optional)}"
                            f" optional headers for {entity_type}"
                        )
            elif required_match > 0:
                partial = required_match / len(required)
                platform_score += 0.2 * partial
                evidence.append(
                    f"{required_match}/{len(required)}"
                    f" required headers for {entity_type}"
                )

        if platform_score > best_score:
            best_score = platform_score
            best_platform = platform
            best_evidence = evidence
            best_entity_types = found_types

    if best_score < 0.5:
        best_platform = "generic"
        best_evidence.append("Low confidence — falling back to generic mapping")

    confidence = min(best_score, 1.0)
    field_mapping = FIELD_MAPPINGS.get(best_platform, {})

    return {
        "platform": best_platform,
        "confidence": round(confidence, 2),
        "evidence": best_evidence,
        "entity_types_found": best_entity_types,
        "row_count": row_count,
        "field_mapping": field_mapping,
    }


def ai_assisted_mapping(headers: list[str], sample_rows: list[dict]) -> dict:
    """Use Claude to infer field mappings when auto-detection fails.

    Sends headers + 3 sample rows. Returns a mapping dict.
    """
    try:
        import anthropic
    except ImportError:
        logger.warning("anthropic not available for AI-assisted mapping")
        return {}

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        logger.warning("ANTHROPIC_API_KEY not set — skipping AI-assisted mapping")
        return {}

    # Limit to 3 sample rows
    samples = sample_rows[:3]

    prompt = (
        "You are a data migration expert. Given these CSV headers and sample rows, "
        "infer what each column represents and map it to these canonical field names:\n"
        "first_name, last_name, email, phone, company, job_title, skills, "
        "location, city, state, salary, external_id, status, source, tags, "
        "linkedin_url, website, industry, employment_type\n\n"
        f"Headers: {headers}\n\n"
        f"Sample rows:\n{json_mod.dumps(samples, indent=2)}\n\n"
        "Return ONLY a JSON object mapping source column names to canonical names. "
        "Only include columns you can confidently map. Example:\n"
        '{"Full Name": "first_name", "E-mail Address": "email"}'
    )

    try:
        client = anthropic.Anthropic(api_key=api_key, max_retries=3)
        response = client.messages.create(
            model="claude-sonnet-4-5-20250514",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.content[0].text
        # Extract JSON from response
        mapping = json_mod.loads(raw)
        if isinstance(mapping, dict):
            return mapping
    except Exception as e:
        logger.warning("AI-assisted mapping failed: %s", e)

    return {}

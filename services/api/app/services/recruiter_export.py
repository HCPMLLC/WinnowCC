"""Export recruiter-sourced candidates to CSV or XLSX."""

import csv
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

COLUMNS = [
    "Name",
    "Headline",
    "Location",
    "Current Company",
    "LinkedIn URL",
    "Skills",
    "Experience Summary",
    "Education Summary",
    "Email",
    "Phone",
    "Source",
    "Date Added",
]


def _flatten_skills(skills: Any) -> str:
    """Convert skills list to comma-separated string."""
    if not skills or not isinstance(skills, list):
        return ""
    names = []
    for s in skills:
        if isinstance(s, str):
            names.append(s)
        elif isinstance(s, dict):
            names.append(s.get("name", ""))
    return ", ".join(n for n in names if n)


def _summarize_experience(experience: Any) -> str:
    """Summarize top 3 experience entries."""
    if not experience or not isinstance(experience, list):
        return ""
    parts = []
    for entry in experience[:3]:
        if not isinstance(entry, dict):
            continue
        title = entry.get("title", "")
        company = entry.get("company", "")
        if title and company:
            parts.append(f"{title} at {company}")
        elif title:
            parts.append(title)
    return "; ".join(parts)


def _summarize_education(education: Any) -> str:
    """Summarize top 3 education entries."""
    if not education or not isinstance(education, list):
        return ""
    parts = []
    for entry in education[:3]:
        if not isinstance(entry, dict):
            continue
        degree = entry.get("degree", "")
        school = entry.get("school") or entry.get("institution", "")
        if degree and school:
            parts.append(f"{degree}, {school}")
        elif school:
            parts.append(school)
    return "; ".join(parts)


def _profile_to_row(profile: Any) -> list[str]:
    """Extract one row of export data from a CandidateProfile."""
    pj = profile.profile_json or {}
    basics = pj.get("basics") or {}

    name = pj.get("name") or basics.get("name") or ""
    if not name:
        first = basics.get("first_name", "")
        last = basics.get("last_name", "")
        name = f"{first} {last}".strip()

    headline = pj.get("headline") or ""
    if not headline:
        titles = basics.get("target_titles") or []
        if titles and isinstance(titles, list):
            headline = titles[0] if isinstance(titles[0], str) else ""

    location = pj.get("location") or basics.get("location") or ""
    current_company = pj.get("current_company") or ""
    linkedin_url = pj.get("linkedin_url") or ""
    skills = _flatten_skills(pj.get("skills") or basics.get("top_skills"))
    experience = _summarize_experience(pj.get("experience"))
    education = _summarize_education(pj.get("education"))
    email = pj.get("email") or basics.get("email") or ""
    phone = pj.get("phone") or basics.get("phone") or ""
    source = pj.get("source") or ""
    date_added = ""
    if profile.updated_at:
        date_added = profile.updated_at.strftime("%Y-%m-%d")

    return [
        name,
        headline,
        location,
        current_company,
        linkedin_url,
        skills,
        experience,
        education,
        email,
        phone,
        source,
        date_added,
    ]


def export_csv(profiles: list[Any]) -> bytes:
    """Generate CSV bytes (utf-8-sig BOM for Excel compatibility)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(COLUMNS)
    for p in profiles:
        writer.writerow(_profile_to_row(p))
    return buf.getvalue().encode("utf-8-sig")


def export_xlsx(profiles: list[Any]) -> bytes:
    """Generate XLSX bytes with bold headers and auto-width columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    # Bold headers
    bold = Font(bold=True)
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Data rows
    for row_idx, p in enumerate(profiles, 2):
        for col_idx, value in enumerate(_profile_to_row(p), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width columns
    for col_idx, header in enumerate(COLUMNS, 1):
        max_len = len(header)
        for row_idx in range(2, len(profiles) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            max_len + 2
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
